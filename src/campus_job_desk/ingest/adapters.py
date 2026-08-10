from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from campus_job_desk.domain.classify import classify_record
from campus_job_desk.domain.enums import IdentityStrength, ParseStatus, SourceKind
from campus_job_desk.domain.normalize import (
    clean_text,
    compound_identity_hint,
    digest,
    extract_official_job_id,
    extract_url,
    is_plausible_official_job_id,
    normalize_url,
    split_values,
)
from campus_job_desk.domain.schemas import (
    CanonicalRecord,
    IdentityHint,
    ParsedRow,
    ParsedSnapshot,
)

from .mapping import UNSTABLE_POSITION_COLUMNS, infer_mapping, normalize_header, validate_mapping

SOURCE_URL_PATTERN = re.compile(r"https?://[^)\s>]+")
SNAPSHOT_TIME_PATTERNS = (
    re.compile(r"快照时间：\s*(.+)$"),
    re.compile(r"文件日期：\s*(.+)$"),
)
DECLARED_COUNT_PATTERNS = (
    re.compile(r"全量记录数：\s*([\d,]+)"),
    re.compile(r"页面显示记录数：\s*([\d,]+)"),
)
COPYRIGHT_MARKERS = ("正版授权", "转售必究")


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = value.strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        try:
            return datetime.strptime(normalized[:10], "%Y-%m-%d")
        except ValueError:
            return None


def _coerce_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return clean_text(value)


def _canonicalize(raw: dict[str, Any], mapping: dict[str, str]) -> CanonicalRecord:
    def get(field: str) -> str:
        return _coerce_cell(raw.get(mapping.get(field, ""), ""))
    announcement_url = extract_url(get("announcement_url"))
    apply_url = extract_url(get("apply_url"))
    title = get("title")
    explicit_job_id_value = get("official_job_id") or None
    explicit_job_id = (
        explicit_job_id_value
        if explicit_job_id_value and is_plausible_official_job_id(explicit_job_id_value)
        else None
    )
    official_job_id = explicit_job_id or extract_official_job_id(title, announcement_url, apply_url)
    source_record_column = mapping.get("source_record_id")
    source_record_id = get("source_record_id") or None
    if source_record_column and normalize_header(source_record_column) in {
        normalize_header(item) for item in UNSTABLE_POSITION_COLUMNS
    }:
        source_record_id = None
    return CanonicalRecord(
        company=get("company"),
        title=title,
        cities=split_values(get("cities")),
        graduation_years=split_values(get("graduation_years")),
        education=split_values(get("education")),
        recruitment_type=get("recruitment_type"),
        industry=get("industry"),
        employer_type=get("employer_type"),
        written_test=get("written_test"),
        published_at=get("published_at"),
        deadline=get("deadline"),
        announcement_url=announcement_url,
        apply_url=apply_url,
        official_job_id=official_job_id,
        source_record_id=source_record_id,
        notes=get("notes"),
    )


def _identity(record: CanonicalRecord) -> IdentityHint:
    if record.source_record_id:
        return IdentityHint(
            value=record.source_record_id,
            strength=IdentityStrength.SOURCE_RECORD_ID,
            is_cross_batch_stable=True,
            evidence="供应商提供稳定记录 ID",
        )
    if record.official_job_id:
        return IdentityHint(
            value=record.official_job_id,
            strength=IdentityStrength.OFFICIAL_JOB_ID,
            is_cross_batch_stable=True,
            evidence="从字段或官方岗位链接提取职位 ID",
        )
    candidate_url = normalize_url(record.apply_url or record.announcement_url)
    if candidate_url:
        return IdentityHint(
            value=candidate_url,
            strength=IdentityStrength.OFFICIAL_URL,
            is_cross_batch_stable=True,
            evidence="规范化岗位或投递链接",
        )
    hint = compound_identity_hint(
        record.company,
        record.title,
        record.cities,
        record.recruitment_type,
        record.graduation_years,
    )
    return IdentityHint(
        value=hint,
        strength=IdentityStrength.COMPOUND_HINT if hint else IdentityStrength.NONE,
        is_cross_batch_stable=False,
        evidence="复合候选键仅用于生成 MATCH_REVIEW，不可自动合并" if hint else "缺少稳定身份信息",
    )


def _rows_to_snapshot(
    *,
    path: Path,
    file_format: str,
    header: list[str],
    raw_rows: list[dict[str, Any]],
    source_name: str,
    source_kind: SourceKind,
    snapshot_at: datetime | None,
    custom_mapping: dict[str, str] | None = None,
    rejected_rows: list[dict[str, Any]] | None = None,
) -> ParsedSnapshot:
    mapping = {**infer_mapping(header), **(custom_mapping or {})}
    mapping_errors = validate_mapping(mapping)
    fatal_errors = [error for error in mapping_errors if error.startswith("缺少")]
    if fatal_errors:
        raise ValueError("；".join(fatal_errors))

    parsed: list[ParsedRow] = []
    for row_number, raw in enumerate(raw_rows, start=1):
        canonical = _canonicalize(raw, mapping)
        errors: list[str] = []
        if not canonical.company:
            errors.append("公司名称为空")
        if not canonical.title:
            errors.append("岗位名称为空")
        parse_status = ParseStatus.REJECTED if len(errors) == 2 else ParseStatus.PARTIAL if errors else ParseStatus.PARSED
        parsed.append(
            ParsedRow(
                row_number=row_number,
                raw_values={str(key): _coerce_cell(value) for key, value in raw.items()},
                canonical=canonical,
                row_hash=digest(raw),
                identity=_identity(canonical),
                kind_prediction=classify_record(canonical),
                parse_status=parse_status,
                errors=errors,
            )
        )

    file_bytes = path.read_bytes()
    return ParsedSnapshot(
        path=path,
        file_name=path.name,
        file_hash=hashlib.sha256(file_bytes).hexdigest(),
        file_format=file_format,
        source_name=source_name,
        source_kind=source_kind,
        snapshot_at=snapshot_at,
        header=header,
        mapping=mapping,
        mapping_version="canonical-v1",
        rows=parsed,
        rejected_rows=(rejected_rows or []) + [
            {"row_number": row.row_number, "errors": row.errors}
            for row in parsed
            if row.parse_status == ParseStatus.REJECTED
        ],
    )


def parse_delimited(
    path: Path,
    *,
    source_name: str,
    source_kind: SourceKind,
    custom_mapping: dict[str, str] | None = None,
) -> ParsedSnapshot:
    content = path.read_text(encoding="utf-8-sig")
    sample = content[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("未找到表头")
    rows = [dict(row) for row in reader]
    return _rows_to_snapshot(
        path=path,
        file_format="tsv" if delimiter == "\t" else "csv",
        header=[clean_text(name) for name in reader.fieldnames],
        raw_rows=rows,
        source_name=source_name,
        source_kind=source_kind,
        snapshot_at=None,
        custom_mapping=custom_mapping,
    )


def parse_xlsx(
    path: Path,
    *,
    source_name: str,
    source_kind: SourceKind,
    custom_mapping: dict[str, str] | None = None,
) -> ParsedSnapshot:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        first_row = next(iterator)
    except StopIteration as exc:
        raise ValueError("工作簿为空") from exc
    header = [_coerce_cell(value) for value in first_row]
    rows = [dict(zip(header, values)) for values in iterator if any(value is not None for value in values)]
    workbook.close()
    return _rows_to_snapshot(
        path=path,
        file_format="xlsx",
        header=header,
        raw_rows=rows,
        source_name=source_name,
        source_kind=source_kind,
        snapshot_at=None,
        custom_mapping=custom_mapping,
    )


def parse_markdown_snapshot(
    path: Path,
    *,
    source_name: str,
    source_kind: SourceKind,
    custom_mapping: dict[str, str] | None = None,
) -> ParsedSnapshot:
    content = path.read_text(encoding="utf-8")
    lines = content.splitlines()
    snapshot_at: datetime | None = None
    declared_count: int | None = None
    for line in lines[:60]:
        for pattern in SNAPSHOT_TIME_PATTERNS:
            match = pattern.search(line)
            if match and snapshot_at is None:
                snapshot_at = _parse_datetime(match.group(1))
        for pattern in DECLARED_COUNT_PATTERNS:
            match = pattern.search(line)
            if match and declared_count is None:
                declared_count = int(match.group(1).replace(",", ""))

    header: list[str] | None = None
    raw_rows: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    inside = False
    for line_number, line in enumerate(lines, start=1):
        cells = line.split("\t") if "\t" in line else []
        if "公司名称" in cells and any(token in cells for token in ("招聘岗位", "岗位名称")):
            header = cells
            inside = True
            continue
        if not inside or line in {"<!-- DATA_START -->", "<!-- DATA_END -->"}:
            continue
        if line.startswith("~~~"):
            break
        if header is None or not line:
            continue
        values = line.split("\t")
        if len(values) != len(header):
            rejected.append(
                {
                    "line_number": line_number,
                    "column_count": len(values),
                    "expected_column_count": len(header),
                }
            )
            continue
        raw_rows.append(dict(zip(header, values)))

    if header is None:
        raise ValueError("Markdown 中未找到 TSV 数据表头")
    if declared_count is not None and len(raw_rows) + len(rejected) != declared_count:
        raise ValueError(
            f"完整性校验失败：声明 {declared_count} 行，读取 {len(raw_rows)} 行，拒绝 {len(rejected)} 行"
        )
    return _rows_to_snapshot(
        path=path,
        file_format="markdown-tsv",
        header=header,
        raw_rows=raw_rows,
        source_name=source_name,
        source_kind=source_kind,
        snapshot_at=snapshot_at,
        custom_mapping=custom_mapping,
        rejected_rows=rejected,
    )


def parse_snapshot(
    path: str | Path,
    *,
    source_name: str,
    source_kind: SourceKind = SourceKind.PAID_TABLE,
    custom_mapping: dict[str, str] | None = None,
) -> ParsedSnapshot:
    resolved = Path(path).expanduser().resolve()
    suffix = resolved.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return parse_delimited(
            resolved,
            source_name=source_name,
            source_kind=source_kind,
            custom_mapping=custom_mapping,
        )
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(
            resolved,
            source_name=source_name,
            source_kind=source_kind,
            custom_mapping=custom_mapping,
        )
    if suffix in {".md", ".markdown", ".txt"}:
        return parse_markdown_snapshot(
            resolved,
            source_name=source_name,
            source_kind=source_kind,
            custom_mapping=custom_mapping,
        )
    raise ValueError(f"不支持的文件格式：{suffix or '无扩展名'}")


def is_non_job_row(row: ParsedRow) -> bool:
    company = row.canonical.company
    return not company or any(marker in company for marker in COPYRIGHT_MARKERS)
